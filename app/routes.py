import math, os, time
from io import BytesIO
from flask import Blueprint, render_template, redirect, url_for, request, send_file, flash, current_app
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from app.models import Incidente, Prediccion, Usuario, Alerta, TipoIncidente, Evidencia, Rol
from app.utils.ml_model import obtener_resumen_riesgo_actual
from app.utils.security import junta_or_admin_required
from app.forms import PerfilForm
from app import db
from datetime import datetime, timedelta, timezone

routes_bp = Blueprint("routes", __name__)


def init_app(app):
    app.register_blueprint(routes_bp)


@routes_bp.route("/")
def index():
    if current_user.is_authenticated:
        return redirect(url_for("routes.dashboard"))
    return render_template("index.html")


@routes_bp.route("/dashboard")
@login_required
def dashboard():
    ahora = datetime.now(timezone.utc)
    hace_30_dias = ahora - timedelta(days=30)

    total_incidentes = Incidente.query.filter(
        Incidente.fecha_hora >= hace_30_dias
    ).count()

    riesgo_actual = obtener_resumen_riesgo_actual()

    altos_30d = Incidente.query.filter(
        Incidente.fecha_hora >= hace_30_dias, Incidente.riesgo_nivel == "alto"
    ).count()

    reportantes_unicos = (
        Incidente.query.with_entities(Incidente.reportado_por)
        .filter(
            Incidente.fecha_hora >= hace_30_dias,
            Incidente.reportado_por.isnot(None),
        )
        .distinct()
        .count()
    )

    ultimos_incidentes = (
        Incidente.query.order_by(Incidente.fecha_hora.desc()).limit(5).all()
    )

    tendencia_datos = []
    for i in range(14):
        dia = ahora - timedelta(days=13 - i)
        inicio = dia.replace(hour=0, minute=0, second=0, microsecond=0)
        fin = dia.replace(hour=23, minute=59, second=59)
        count = Incidente.query.filter(
            Incidente.fecha_hora >= inicio, Incidente.fecha_hora <= fin
        ).count()
        tendencia_datos.append({"fecha": dia.strftime("%d/%m"), "total": count})

    tipos_datos = []
    tipos_query = (
        db.session.query(TipoIncidente.nombre, db.func.count(Incidente.id))
        .join(Incidente, TipoIncidente.id == Incidente.tipo_incidente_id)
        .filter(Incidente.fecha_hora >= hace_30_dias)
        .group_by(TipoIncidente.nombre)
        .all()
    )
    for t, c in tipos_query:
        tipos_datos.append({"tipo": t, "total": c})

    alertas_no_leidas = Alerta.query.filter_by(
        usuario_id=current_user.id, leida=False
    ).count()

    total_incidentes_all = Incidente.query.count()
    total_vecinos = Usuario.query.join(Rol).filter(Rol.nombre == "vecino").count()
    total_alertas = Alerta.query.count()
    total_usuarios = Usuario.query.count()
    verificados_30d = Incidente.query.filter(
        Incidente.fecha_hora >= hace_30_dias, Incidente.verificado == True
    ).count()

    safety_score = 0
    if total_incidentes > 0:
        ratio = altos_30d / total_incidentes
        safety_score = max(10, min(100, int(100 - (ratio * 100) - (riesgo_actual["zonas_altas"] * 2))))
    elif riesgo_actual["zonas_altas"] > 0:
        safety_score = max(10, 80 - riesgo_actual["zonas_altas"] * 5)
    else:
        safety_score = 85

    template = "dashboard_admin.html" if current_user.es_admin() else "dashboard_vecino.html"
    if current_user.es_junta() and not current_user.es_admin():
        template = "dashboard_admin.html"

    mis_alertas = Alerta.query.filter_by(usuario_id=current_user.id).order_by(Alerta.fecha_envio.desc()).limit(5).all()

    return render_template(
        template,
        total_incidentes=total_incidentes,
        total_incidentes_all=total_incidentes_all,
        total_vecinos=total_vecinos,
        total_alertas=total_alertas,
        riesgo_actual=riesgo_actual,
        altos_30d=altos_30d,
        reportantes_unicos=reportantes_unicos,
        ultimos_incidentes=ultimos_incidentes,
        tendencia_datos=tendencia_datos,
        tipos_datos=tipos_datos,
        alertas_no_leidas=alertas_no_leidas,
        ahora=ahora,
        total_usuarios=total_usuarios,
        verificados_30d=verificados_30d,
        safety_score=safety_score,
        mis_alertas=mis_alertas,
    )


@routes_bp.route("/reportes")
@login_required
def reportes():
    ahora = datetime.now(timezone.utc)
    fecha_hasta = ahora.strftime("%Y-%m-%d")
    fecha_desde = (ahora - timedelta(days=30)).strftime("%Y-%m-%d")
    tipos = TipoIncidente.query.filter_by(activo=True).all()
    return render_template("reportes_export.html", ahora=ahora, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, tipos=tipos)


@routes_bp.route("/perfil", methods=["GET", "POST"])
@login_required
def perfil():
    form = PerfilForm(obj=current_user)
    if form.validate_on_submit():
        email = form.email.data.strip().lower()
        if email != current_user.email and Usuario.query.filter_by(email=email).first():
            flash("Ese email ya está registrado por otro usuario.", "error")
            return render_template("perfil.html", form=form)

        current_user.nombre = form.nombre.data.strip()
        current_user.email = email
        current_user.telefono = form.telefono.data.strip() if form.telefono.data else None
        current_user.barrio_id = form.barrio_id.data

        if form.foto.data and form.foto.data.filename:
            f = form.foto.data
            ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else "jpg"
            filename = f"profile_{int(time.time())}_{secure_filename(f.filename)}"
            profiles_dir = os.path.join(current_app.root_path, "static", "uploads", "profiles")
            os.makedirs(profiles_dir, exist_ok=True)
            f.save(os.path.join(profiles_dir, filename))
            current_user.foto_url = f"uploads/profiles/{filename}"

        if form.password.data:
            if form.password.data != form.confirm_password.data:
                flash("Las contraseñas no coinciden.", "error")
                return render_template("perfil.html", form=form)
            current_user.set_password(form.password.data)

        db.session.commit()
        flash("Perfil actualizado correctamente.", "success")
        return redirect(url_for("routes.perfil"))

    mis_incidentes = Incidente.query.filter_by(reportado_por=current_user.id).order_by(Incidente.fecha_hora.desc()).limit(5).all()
    mis_alertas = Alerta.query.filter_by(usuario_id=current_user.id).order_by(Alerta.fecha_envio.desc()).limit(5).all()
    total_reportados = Incidente.query.filter_by(reportado_por=current_user.id).count()
    verificados = Incidente.query.filter_by(reportado_por=current_user.id, verificado=True).count()

    return render_template("perfil.html", form=form,
        mis_incidentes=mis_incidentes, mis_alertas=mis_alertas,
        total_reportados=total_reportados, verificados=verificados,
        pendientes=total_reportados - verificados)


def _filtrar_incidentes():
    tipo = request.args.get("tipo")
    riesgo = request.args.get("riesgo")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")
    query = Incidente.query.join(TipoIncidente)
    if tipo:
        query = query.filter(TipoIncidente.nombre == tipo)
    if riesgo:
        query = query.filter(Incidente.riesgo_nivel == riesgo)
    if fecha_desde:
        query = query.filter(Incidente.fecha_hora >= datetime.fromisoformat(fecha_desde))
    if fecha_hasta:
        query = query.filter(Incidente.fecha_hora <= datetime.fromisoformat(fecha_hasta))
    return query.order_by(Incidente.fecha_hora.desc()).all()


@routes_bp.route("/reportes/exportar/excel")
@login_required
@junta_or_admin_required
def exportar_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    incidentes = _filtrar_incidentes()
    from app.models import Evidencia as EvidenciaModel

    wb = Workbook()
    ws = wb.active
    ws.title = "Incidentes"

    headers = ["#", "Tipo", "Descripción", "Fecha", "Latitud", "Longitud", "Riesgo", "Estado", "Dirección", "Evidencias"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    riesgo_fills = {
        "alto": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        "medio": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "bajo": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
    }

    for i, inc in enumerate(incidentes, 1):
        row = i + 1
        ev_count = EvidenciaModel.query.filter_by(incidente_id=inc.id).count()
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=inc.tipo)
        ws.cell(row=row, column=3, value=inc.descripcion or "—")
        ws.cell(row=row, column=4, value=inc.fecha_hora.strftime("%d/%m/%Y %H:%M") if inc.fecha_hora else "—")
        ws.cell(row=row, column=5, value=round(inc.latitud, 6))
        ws.cell(row=row, column=6, value=round(inc.longitud, 6))
        ws.cell(row=row, column=7, value=inc.riesgo_nivel)
        ws.cell(row=row, column=8, value="Verificado" if inc.verificado else "Pendiente")
        ws.cell(row=row, column=9, value=inc.direccion or "—")
        ws.cell(row=row, column=10, value="%d foto(s)" % ev_count if ev_count else "—")

        fill = riesgo_fills.get(inc.riesgo_nivel)
        if fill:
            ws.cell(row=row, column=7).fill = fill
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = thin_border

    for col in range(1, 11):
        ws.column_dimensions[chr(64 + col)].width = max(
            12, len(str(ws.cell(row=1, column=col).value or "")) + 2
        )

    ws2 = wb.create_sheet("Evidencias")
    ev_headers = ["ID Evidencia", "ID Incidente", "Tipo", "Archivo", "Fecha Subida"]
    for col, h in enumerate(ev_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    evidencias = EvidenciaModel.query.order_by(EvidenciaModel.incidente_id).all()
    for i, ev in enumerate(evidencias, 1):
        ws2.cell(row=i+1, column=1, value=ev.id)
        ws2.cell(row=i+1, column=2, value=ev.incidente_id)
        ws2.cell(row=i+1, column=3, value=ev.tipo)
        ws2.cell(row=i+1, column=4, value=ev.archivo_url)
        ws2.cell(row=i+1, column=5, value=ev.fecha_subida.strftime("%d/%m/%Y %H:%M") if ev.fecha_subida else "—")
        for col in range(1, 6):
            ws2.cell(row=i+1, column=col).border = thin_border

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 50
    ws2.column_dimensions["E"].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"incidentes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )


@routes_bp.route("/reportes/exportar/pdf")
@login_required
@junta_or_admin_required
def exportar_pdf():
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    incidentes = _filtrar_incidentes()

    output = BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title2", parent=styles["Title"], fontSize=18, spaceAfter=6, textColor=HexColor("#1B4332"))
    subtitle_style = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=10, textColor=HexColor("#6c757d"), spaceAfter=20)

    elements = []
    elements.append(Paragraph("SEGURIDAD VECINAL — Reporte de Incidentes", title_style))
    elements.append(Paragraph("Generado: %s | Total: %d incidentes" % (
        datetime.now().strftime("%d/%m/%Y %H:%M"), len(incidentes)
    ), subtitle_style))
    elements.append(Spacer(1, 0.5*cm))

    headers = ["#", "Tipo", "Descripción", "Fecha", "Lat/Lng", "Riesgo", "Estado", "Evidencias"]
    data = [headers]
    from app.models import Evidencia as EvidenciaModel

    for i, inc in enumerate(incidentes, 1):
        desc = (inc.descripcion[:60] + "..") if inc.descripcion and len(inc.descripcion) > 60 else (inc.descripcion or "—")
        ev_count = EvidenciaModel.query.filter_by(incidente_id=inc.id).count()
        data.append([
            str(i),
            inc.tipo,
            desc,
            inc.fecha_hora.strftime("%d/%m/%Y %H:%M") if inc.fecha_hora else "—",
            "%.4f, %.4f" % (inc.latitud, inc.longitud),
            inc.riesgo_nivel,
            "Verificado" if inc.verificado else "Pendiente",
            "%d foto(s)" % ev_count if ev_count else "—",
        ])

    col_widths = [1.2*cm, 2.5*cm, 5*cm, 3*cm, 3.5*cm, 1.8*cm, 2.2*cm, 2*cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)

    riesgo_colors = {"alto": HexColor("#F8D7DA"), "medio": HexColor("#FFF3CD"), "bajo": HexColor("#D4EDDA")}

    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#1B4332")),
        ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#FFFFFF")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (2, 1), (2, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#DEE2E6")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [HexColor("#FFFFFF"), HexColor("#F8F9FA")]),
    ]
    for i_row in range(1, len(data)):
        riesgo_val = data[i_row][5]
        color = riesgo_colors.get(riesgo_val)
        if color:
            style_cmds.append(("BACKGROUND", (5, i_row), (5, i_row), color))

    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    doc.build(elements)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"incidentes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
    )


@routes_bp.route("/reportes/exportar/csv")
@login_required
@junta_or_admin_required
def exportar_csv():
    import csv as csv_module
    from app.models import Evidencia as EvidenciaModel

    incidentes = _filtrar_incidentes()
    output = BytesIO()
    text_output = output  # BytesIO for csv
    import io
    text_output = io.StringIO()
    writer = csv_module.writer(text_output)
    writer.writerow(["ID", "Tipo", "Descripcion", "Fecha", "Latitud", "Longitud", "Riesgo", "Estado", "Direccion", "Evidencias"])
    for inc in incidentes:
        ev_count = EvidenciaModel.query.filter_by(incidente_id=inc.id).count()
        writer.writerow([
            inc.id, inc.tipo, inc.descripcion or "",
            inc.fecha_hora.strftime("%Y-%m-%d %H:%M") if inc.fecha_hora else "",
            round(inc.latitud, 6), round(inc.longitud, 6),
            inc.riesgo_nivel, "Verificado" if inc.verificado else "Pendiente",
            inc.direccion or "", ev_count,
        ])
    bytes_output = BytesIO(text_output.getvalue().encode("utf-8-sig"))
    return send_file(
        bytes_output,
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"incidentes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
    )
